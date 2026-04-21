"""
GIK Institute Timetable Exporter
PDF  → GIK-style Room × Timeslot grid, one page per day, cells show "CODE SEC"
Excel→ Per-day grid sheets (same layout) + All Sessions list + Summary
"""

import os
from typing import List, Dict
from core.models import Timetable, Assignment

DAY_ORDER = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

# Regular Mon-Thu slots
SLOTS_REGULAR = [
    ('08:00', '08:50'), ('09:00', '09:50'),
    ('10:30', '11:20'), ('11:30', '12:20'), ('12:30', '13:20'),
    ('14:30', '15:20'), ('15:30', '16:20'), ('16:30', '17:20'),
]
# Friday slots
SLOTS_FRIDAY = [
    ('08:00', '08:50'), ('09:00', '09:50'),
    ('10:00', '10:50'), ('11:00', '11:50'), ('12:00', '12:50'),
    ('14:30', '15:20'), ('15:30', '16:20'), ('16:30', '17:20'),
]

# Preferred room display order (matches official GIK timetable)
ROOM_ORDER = [
    'CS LH1', 'CS LH2', 'CS LH3',
    'EE LH4', 'EE LH5', 'EE LH6', 'EE Main',
    'ES LH1', 'ES LH2', 'ES LH3', 'ES Main',
    'FES - PH Lab', 'PC Lab',
    'AcB LH1',  'AcB LH2',  'AcB LH3',  'AcB LH4',  'AcB LH5',
    'AcB LH6',  'AcB LH7',  'AcB LH8',  'AcB LH9',  'AcB LH10',
    'AcB LH11', 'AcB LH12',
    'AcB Main1', 'AcB Main2', 'AcB Main3',
    'BB LH2', 'BB EH1', 'BB EH2', 'BB EH3', 'BB EH4', 'BB Main', 'BB PC Lab',
    'ME LH1', 'ME LH2', 'ME LH3', 'ME Main',
    'TBA',
    'MCE LH1', 'MCE LH2', 'MCE LH3', 'MCE LH4',
    'FCME - MM Lab', 'FCME - CH Lab', 'MCE Main',
    'FME Lab',
    'FCSE - SE Lab', 'FES - SE Lab',
    'ACB - AI Lab', 'ACB - CYS Lab', 'ACB - DA Lab',
]

DAY_COLOURS = {
    'Monday':    'DBEAFE', 'Tuesday':  'D1FAE5',
    'Wednesday': 'FEF9C3', 'Thursday': 'FFE4E6', 'Friday': 'EDE9FE',
}
LAB_BG    = 'FFF3CD'
HDR_BG    = 'C8D8F0'
ROOM_BG   = 'EFF6FF'


# ── Helpers ───────────────────────────────────────────────────────────────────

def _slots_for_day(day: str):
    return SLOTS_FRIDAY if day == 'Friday' else SLOTS_REGULAR


def _cell_text(a: Assignment) -> str:
    return f"{a.course.code} {a.course.section}".strip()


def _build_grid(day_assignments: List[Assignment], day: str):
    """
    Returns:
      grid[room_name][start_time] = Assignment
      ordered_rooms  – rooms sorted by ROOM_ORDER
      slot_starts    – list of start-time strings in display order
    """
    slot_starts = [s for s, _ in _slots_for_day(day)]
    grid: Dict[str, Dict[str, Assignment]] = {}
    for a in day_assignments:
        rname = a.room.room_name
        grid.setdefault(rname, {})[a.timeslot.start_time] = a

    used = set(grid.keys())
    ordered = [r for r in ROOM_ORDER if r in used]
    ordered += sorted(r for r in used if r not in ordered)
    return grid, ordered, slot_starts


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def export_excel(timetable: Timetable, output_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)          # remove the blank default sheet

    thin   = Side(border_style='thin',   color='CCCCCC')
    medium = Side(border_style='medium', color='888888')
    b_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_med  = Border(left=medium, right=medium, top=medium, bottom=medium)

    def fill(hex_str):
        return PatternFill('solid', fgColor=hex_str)

    hdr_font  = Font(name='Calibri', bold=True, size=9,  color='1E3A5F')
    room_font = Font(name='Calibri', bold=True, size=8,  color='1E3A5F')
    cell_font = Font(name='Calibri', size=9)
    lab_font  = Font(name='Calibri', bold=True, size=9,  color='7C4700')
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_al   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ── Grid sheet per day (these are the main sheets users care about) ────
    for day in DAY_ORDER:
        day_asns = [a for a in timetable.assignments if a.timeslot.day == day]
        if not day_asns:
            continue

        ws = wb.create_sheet(day[:3])
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'B3'      # freeze room col + title/header rows

        slots       = _slots_for_day(day)
        slot_starts = [s for s, _ in slots]
        N           = len(slots)          # always 8
        grid, rooms, _ = _build_grid(day_asns, day)
        day_fill    = fill(DAY_COLOURS[day])
        lab_fill    = fill(LAB_BG)
        hdr_fill    = fill(HDR_BG)
        room_fill   = fill(ROOM_BG)

        # Col widths: A=room, B..I=slots
        ws.column_dimensions['A'].width = 20
        for ci in range(2, N + 2):
            ws.column_dimensions[get_column_letter(ci)].width = 13

        # ── Row 1: title spanning all columns ─────────────────────────────
        last_col = get_column_letter(N + 1)
        ws.merge_cells(f'A1:{last_col}1')
        tc = ws['A1']
        tc.value     = f'GIK Institute  —  Time Table  Fall 2025   ({day})'
        tc.font      = Font(name='Calibri', bold=True, size=12, color='1E3A5F')
        tc.fill      = hdr_fill
        tc.alignment = center
        ws.row_dimensions[1].height = 24

        # ── Row 2: slot headers ────────────────────────────────────────────
        ws.cell(2, 1, 'Room').font      = hdr_font
        ws.cell(2, 1).fill      = hdr_fill
        ws.cell(2, 1).alignment = center
        ws.cell(2, 1).border    = b_thin
        ws.row_dimensions[2].height = 28

        for ci, (st, et) in enumerate(slots, start=2):
            c = ws.cell(2, ci, f'{st}–{et}')
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = b_thin

        # ── Data rows ──────────────────────────────────────────────────────
        for ri, rname in enumerate(rooms, start=3):
            ws.row_dimensions[ri].height = 20
            rc = ws.cell(ri, 1, rname)
            rc.font = room_font; rc.fill = room_fill
            rc.alignment = left_al; rc.border = b_thin

            room_slots = grid.get(rname, {})
            ci = 2                       # openpyxl column (2 = first slot col)
            occupied = set()             # columns already merged into a lab span

            while ci <= N + 1:
                if ci in occupied:
                    ws.cell(ri, ci).fill   = lab_fill
                    ws.cell(ri, ci).border = b_thin
                    ci += 1
                    continue

                st = slot_starts[ci - 2]     # ci=2 → index 0
                a  = room_slots.get(st)

                if a is None:
                    c = ws.cell(ri, ci)
                    c.fill = day_fill; c.border = b_thin
                    ci += 1
                    continue

                text   = _cell_text(a)
                is_lab = a.course.is_lab

                if is_lab:
                    end_ci = min(ci + 2, N + 1)
                    ws.merge_cells(
                        start_row=ri, start_column=ci,
                        end_row=ri,   end_column=end_ci
                    )
                    c = ws.cell(ri, ci, text)
                    c.font = lab_font; c.fill = lab_fill
                    c.alignment = center; c.border = b_med
                    for sc in range(ci + 1, end_ci + 1):
                        occupied.add(sc)
                    ci = end_ci + 1
                else:
                    c = ws.cell(ri, ci, text)
                    c.font = cell_font; c.fill = day_fill
                    c.alignment = center; c.border = b_thin
                    ci += 1

    # ── All Sessions list sheet ────────────────────────────────────────────
    _write_list_sheet(wb, timetable)

    # ── Summary sheet — insert at position 0 so it's the first tab ────────
    _write_summary_sheet(wb, timetable)
    # Move Summary to front
    wb.move_sheet('Summary', offset=-len(wb.sheetnames) + 1)

    wb.save(output_path)
    return output_path


def _write_list_sheet(wb, timetable):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin   = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfont  = Font(name='Calibri', bold=True, size=9, color='1E3A5F')
    dfont  = Font(name='Calibri', size=9)
    hfill  = PatternFill('solid', fgColor=HDR_BG)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    COLS = [
        ('Day', 11), ('Start Time', 9), ('End Time', 9),
        ('Course Code', 12), ('Section', 7), ('Course Title', 34),
        ('Type', 8), ('Credit Hours', 8), ('Instructor', 24),
        ('Program', 10), ('Location', 24), ('Building', 12),
    ]

    ws = wb.create_sheet('All Sessions')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    for ci, (name, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        c = ws.cell(1, ci, name)
        c.font = hfont; c.fill = hfill
        c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 24

    rows = sorted(
        timetable.to_dict_list(),
        key=lambda r: (
            DAY_ORDER.index(r['Day']) if r['Day'] in DAY_ORDER else 9,
            r['Start Time']
        )
    )
    day_fills = {d: PatternFill('solid', fgColor=DAY_COLOURS[d]) for d in DAY_ORDER}
    lab_fill  = PatternFill('solid', fgColor=LAB_BG)

    for ri, row in enumerate(rows, 2):
        f = lab_fill if row['Type'] == 'Lab' else day_fills.get(row['Day'])
        for ci, (name, _) in enumerate(COLS, 1):
            c = ws.cell(ri, ci, row.get(name, ''))
            c.font = dfont; c.alignment = left; c.border = border
            if f:
                c.fill = f
        ws.row_dimensions[ri].height = 16


def _write_summary_sheet(wb, timetable):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin   = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    tfont  = Font(name='Calibri', bold=True, size=13, color='1E3A5F')
    hfont  = Font(name='Calibri', bold=True, size=9,  color='1E3A5F')
    dfont  = Font(name='Calibri', size=9)
    hfill  = PatternFill('solid', fgColor=HDR_BG)
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    ws = wb.create_sheet('Summary')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16

    ws.merge_cells('A1:B1')
    ws['A1'] = 'GIK Timetable — Summary  (Fall 2025)'
    ws['A1'].font = tfont; ws['A1'].fill = hfill
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    def kv(row, key, val, bg=None):
        c1 = ws.cell(row, 1, key)
        c2 = ws.cell(row, 2, val)
        c1.font = hfont; c2.font = dfont
        c1.alignment = left; c2.alignment = center
        c1.border = border; c2.border = border
        if bg:
            f = PatternFill('solid', fgColor=bg)
            c1.fill = f; c2.fill = f
        ws.row_dimensions[row].height = 18

    labs   = sum(1 for a in timetable.assignments if a.course.is_lab)
    lects  = timetable.total_scheduled - labs

    r = 3
    kv(r, 'Total Sessions',         timetable.total_scheduled); r += 1
    kv(r, '  └ Lecture Sessions',   lects);                     r += 1
    kv(r, '  └ Lab Blocks (3 hr)',  labs);                      r += 2
    kv(r, 'Sessions by Day', '')
    r += 1
    for day in DAY_ORDER:
        kv(r, f'  {day}', len(timetable.get_by_day(day)), DAY_COLOURS[day])
        r += 1
    r += 1
    progs = sorted(set(a.course.program for a in timetable.assignments))
    kv(r, 'Sessions by Program', '')
    r += 1
    for p in progs:
        kv(r, f'  {p}', len(timetable.get_by_program(p)))
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF EXPORT  —  GIK-style grid
# ═══════════════════════════════════════════════════════════════════════════════

def export_pdf(timetable: Timetable, output_path: str) -> str:
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, PageBreak
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    def hx(s):
        return colors.HexColor(f'#{s}')

    day_cl  = {d: hx(DAY_COLOURS[d]) for d in DAY_ORDER}
    hdr_bg  = hx(HDR_BG)
    room_bg = hx(ROOM_BG)
    lab_bg  = hx(LAB_BG)
    grid_ln = hx('CCCCCC')
    dark    = hx('1E3A5F')
    lab_fg  = hx('7C4700')
    white   = colors.white

    SS = getSampleStyleSheet()
    def ps(name, **kw):
        return ParagraphStyle(name, parent=SS['Normal'], **kw)

    title_s = ps('T',  fontSize=12, leading=15, alignment=TA_CENTER,
                 textColor=dark, fontName='Helvetica-Bold')
    day_s   = ps('D',  fontSize=10, leading=13, alignment=TA_CENTER,
                 textColor=dark, fontName='Helvetica-Bold')
    hdr_s   = ps('H',  fontSize=6,  leading=7.5, alignment=TA_CENTER,
                 textColor=dark, fontName='Helvetica-Bold')
    room_s  = ps('R',  fontSize=6.5,leading=8,   alignment=TA_LEFT,
                 textColor=dark, fontName='Helvetica-Bold')
    cell_s  = ps('C',  fontSize=7,  leading=8.5, alignment=TA_CENTER,
                 textColor=hx('0F172A'))
    lab_s   = ps('L',  fontSize=7,  leading=8.5, alignment=TA_CENTER,
                 textColor=lab_fg, fontName='Helvetica-Bold')

    # A3 landscape usable width
    L_MAR, R_MAR = 6*mm, 6*mm
    USABLE = 420*mm - L_MAR - R_MAR    # ≈ 408 mm
    ROOM_W = 28*mm
    N_SLOTS = 8
    SLOT_W  = (USABLE - ROOM_W) / N_SLOTS   # ≈ 47.5 mm

    col_widths = [ROOM_W] + [SLOT_W] * N_SLOTS

    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A3),
        leftMargin=L_MAR, rightMargin=R_MAR,
        topMargin=8*mm, bottomMargin=8*mm,
    )

    story = []
    story.append(Paragraph('GIK Institute  —  Time Table  Fall 2025', title_s))
    story.append(Spacer(1, 4*mm))

    for day_idx, day in enumerate(DAY_ORDER):
        day_asns = [a for a in timetable.assignments if a.timeslot.day == day]
        if not day_asns:
            continue

        slots       = _slots_for_day(day)
        slot_starts = [s for s, _ in slots]
        grid, rooms, _ = _build_grid(day_asns, day)
        bg = day_cl[day]

        # ── Header row ────────────────────────────────────────────────────
        hdr_row = [Paragraph('Room', hdr_s)]
        for st, et in slots:
            hdr_row.append(Paragraph(f'{st}\n{et}', hdr_s))

        table_data = [hdr_row]
        # Collect style commands that apply to individual cells/spans
        span_cmds  = []
        bg_cmds    = []

        # ── Data rows ─────────────────────────────────────────────────────
        for ri, rname in enumerate(rooms, start=1):   # ri=1 because row 0 = header
            data_row = [Paragraph(rname, room_s)]
            # ci = slot index 0..7 (0-based inside the slot array)
            ci = 0
            skip = set()   # slot indices already covered by a lab span

            while ci < N_SLOTS:
                if ci in skip:
                    data_row.append('')
                    ci += 1
                    continue

                st = slot_starts[ci]
                a  = grid.get(rname, {}).get(st)

                if a is None:
                    data_row.append('')
                    ci += 1
                    continue

                text   = _cell_text(a)
                is_lab = a.course.is_lab

                if is_lab:
                    span_end = min(ci + 2, N_SLOTS - 1)
                    data_row.append(Paragraph(text, lab_s))
                    for sc in range(ci + 1, span_end + 1):
                        data_row.append('')
                        skip.add(sc)
                    # Table column indices (0-indexed):
                    #   col 0 = Room col
                    #   col 1..8 = slot cols  → slot ci → table col ci+1
                    tc_start = ci + 1
                    tc_end   = span_end + 1
                    span_cmds.append(('SPAN',       (tc_start, ri), (tc_end, ri)))
                    bg_cmds.append(  ('BACKGROUND', (tc_start, ri), (tc_end, ri), lab_bg))
                    bg_cmds.append(  ('BOX',        (tc_start, ri), (tc_end, ri),
                                      1.0, hx('D97706')))
                    ci = span_end + 1
                else:
                    data_row.append(Paragraph(text, cell_s))
                    ci += 1

            table_data.append(data_row)

        # ── Base style ────────────────────────────────────────────────────
        n_rows = len(table_data)
        style_cmds = [
            # Header row
            ('BACKGROUND',    (0, 0), (-1, 0),  hdr_bg),
            ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0),  6),
            ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('LINEBELOW',     (0, 0), (-1, 0),  0.8, dark),
            # Room column
            ('BACKGROUND',    (0, 1), (0, -1),  room_bg),
            ('FONTNAME',      (0, 1), (0, -1),  'Helvetica-Bold'),
            ('FONTSIZE',      (0, 1), (0, -1),  6.5),
            ('ALIGN',         (0, 1), (0, -1),  'LEFT'),
            ('LINEAFTER',     (0, 0), (0, -1),  0.8, dark),
            # Data cells default background = day colour
            ('BACKGROUND',    (1, 1), (-1, -1), bg),
            # Data cell text
            ('FONTNAME',      (1, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',      (1, 1), (-1, -1), 7),
            ('ALIGN',         (1, 1), (-1, -1), 'CENTER'),
            # Grid
            ('GRID',          (0, 0), (-1, -1), 0.3, grid_ln),
            # Padding
            ('TOPPADDING',    (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING',   (0, 0), (-1, -1), 2),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 2),
        ]
        # Apply lab spans AFTER setting default bg (so they override it)
        style_cmds += bg_cmds
        style_cmds += span_cmds

        row_heights = [20] + [11] * (n_rows - 1)

        t = Table(table_data, colWidths=col_widths,
                  rowHeights=row_heights, repeatRows=1)
        t.setStyle(TableStyle(style_cmds))

        story.append(Paragraph(f'◆  {day}', day_s))
        story.append(Spacer(1, 2*mm))
        story.append(t)

        if day_idx < len(DAY_ORDER) - 1:
            story.append(PageBreak())

    doc.build(story)
    return output_path
